import csv

from openpyxl import load_workbook


def reformat(file_name):
    """Reformat table.xlsx to a file.csv with only necessary columns"""

    wb = load_workbook(filename=f'files/{file_name}.xlsx')
    ws = wb.active
    ws.unmerge_cells('A1:Q1')
    ws.delete_rows(1)

    ws.delete_cols(1, 4)
    ws.delete_cols(5, 4)
    ws.delete_cols(6, 1)
    ws.delete_cols(7, 2)

    rows = ws.max_row
    print(f'Quantity of records: {rows}.')

    ws['A1'].value = 'spec_code_id'
    ws['B1'].value = 'financing_type'
    ws['C1'].value = 'originals'
    ws['D1'].value = 'avg_marks'
    ws['E1'].value = 'grade'
    ws['F1'].value = 'certificate_number'

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1, max_row=rows):
        for cell in row:
            if 'Техническая эксплуатация' in cell.value:
                cell.value = '13.02.11'
            elif 'Графический' in cell.value:
                cell.value = '54.01.20'
            elif 'Информационные' in cell.value:
                cell.value = '09.02.07'
            elif 'Компьютерные' in cell.value:
                cell.value = '09.02.01'
            elif 'Сетевое' in cell.value:
                cell.value = '09.02.06'
            elif 'Обеспечение' in cell.value:
                cell.value = '10.02.05'
            elif 'Электр' in cell.value:
                cell.value = '13.01.14'

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=rows):
        for cell in row:
            if cell.value == 'Бюджет':
                cell.value = True
            else:
                cell.value = False

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=rows):
        for cell in row:
            if cell.value == 'Сдан':
                cell.value = True
            else:
                cell.value = False

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, max_row=rows):
        for cell in row:
            if cell.value == 'Аттестат об основном общем образовании':
                cell.value = True
            else:
                cell.value = False

    wb.save(f'files/{file_name}_m.xlsx')
    wb.close()

    with open(f'files/{file_name}.csv', 'w', newline="") as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])

    print('File was reformatted successfully.')


if __name__ == '__main__':
    reformat(input("Enter date of the file - the file_name before '.xlsx'"))
